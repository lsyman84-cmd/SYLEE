from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

# ─── 한글 폰트 등록 (맑은 고딕) ──────────────────────────────────────────────
_KR  = "Helvetica"       # fallback (영문)
_KRB = "Helvetica-Bold"  # fallback bold

try:
    _MALGUN_PATH     = r"C:\Windows\Fonts\malgun.ttf"
    _MALGUN_BD_PATH  = r"C:\Windows\Fonts\malgunbd.ttf"
    if Path(_MALGUN_PATH).exists():
        pdfmetrics.registerFont(TTFont("Malgun", _MALGUN_PATH))
        _KR = "Malgun"
    if Path(_MALGUN_BD_PATH).exists():
        pdfmetrics.registerFont(TTFont("MalgunBold", _MALGUN_BD_PATH))
        _KRB = "MalgunBold"
    if _KR == "Malgun" and _KRB == "MalgunBold":
        registerFontFamily("Malgun", normal="Malgun", bold="MalgunBold",
                           italic="Malgun", boldItalic="MalgunBold")
except Exception:
    pass

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# ─── 스타일 상수 ─────────────────────────────────────────────────────────────
_NAVY      = "1F3864"
_BLUE      = "2E75B6"
_GREEN     = "C6EFCE"
_YELLOW    = "FFEB9C"
_RED_LIGHT = "FFC7CE"
_GREY_ROW  = "F2F2F2"

_RISK_HEX  = {"HIGH": "FF4444", "MEDIUM": "FFA500", "LOW": "44BB44"}
_RISK_FILL = {"HIGH": PatternFill("solid", fgColor="FF4444"),
              "MEDIUM": PatternFill("solid", fgColor="FFA500"),
              "LOW":  PatternFill("solid", fgColor="44BB44")}
_WHITE_FONT = Font(bold=True, color="FFFFFF", size=9)

_HDR_FILL  = PatternFill("solid", fgColor=_NAVY)
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=9)
_SUB_FILL  = PatternFill("solid", fgColor=_BLUE)
_SUB_FONT  = Font(bold=True, color="FFFFFF", size=9)
_THIN      = Side(style="thin")
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP      = Alignment(wrap_text=True, vertical="top")


# ─── 공개 API ────────────────────────────────────────────────────────────────

def generate_report(
    conditions: dict,
    candidates_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
    risk_df: pd.DataFrame,
    ai_result: dict,
    filter_log: list,
) -> dict:
    OUTPUT_DIR.mkdir(exist_ok=True)
    today = date.today().strftime("%Y%m%d")
    stem = f"equipment_selection_report_{today}"
    excel_path = OUTPUT_DIR / f"{stem}.xlsx"
    pdf_path   = OUTPUT_DIR / f"{stem}.pdf"

    _write_excel(excel_path, conditions, candidates_df, comparison_df, risk_df, ai_result)
    _write_pdf(pdf_path, conditions, comparison_df, risk_df, ai_result)

    print(f"[SUCCESS] Excel: {excel_path.name}")
    print(f"[SUCCESS] PDF:   {pdf_path.name}")
    return {"excel": str(excel_path), "pdf": str(pdf_path)}


# ─── Excel 공통 헬퍼 ─────────────────────────────────────────────────────────

def _hdr(ws, r, c, val, fill=None, font=None, span=1, height=None):
    cell = ws.cell(r, c, val)
    cell.fill = fill or _HDR_FILL
    cell.font = font or _HDR_FONT
    cell.alignment = _CENTER
    cell.border = _BORDER
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    if height:
        ws.row_dimensions[r].height = height
    return cell


def _cell(ws, r, c, val, bold=False, wrap=False, align="left", fill=None, font_color=None):
    cell = ws.cell(r, c, val)
    cell.border = _BORDER
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap
    )
    if bold or font_color:
        cell.font = Font(bold=bold, color=font_color or "000000")
    if fill:
        cell.fill = fill
    return cell


def _auto_width(ws, min_w=8, max_w=45):
    for col in ws.columns:
        w = min_w
        for cell in col:
            try:
                w = max(w, min(max_w, len(str(cell.value or "")) + 3))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = w


def _score_fill(val):
    if not isinstance(val, (int, float)):
        return None
    if val >= 0.65:
        return PatternFill("solid", fgColor=_GREEN)
    if val >= 0.40:
        return PatternFill("solid", fgColor=_YELLOW)
    return PatternFill("solid", fgColor=_RED_LIGHT)


def _margin_pct(equip_val, input_val):
    try:
        if equip_val and input_val and float(input_val) > 0:
            return (float(equip_val) - float(input_val)) / float(input_val) * 100
    except Exception:
        pass
    return None


def _fmt(val, decimals=1):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "—"
    if isinstance(val, float):
        return f"{val:.{decimals}f}"
    return str(val)


# ─── Sheet 1: 핵심 요약 ──────────────────────────────────────────────────────

def _sheet_summary(wb, conditions, comparison_df, risk_df, ai_result):
    ws = wb.create_sheet("1. 핵심 요약")

    # ── 제목 ──
    title = ws.cell(1, 1, "MECHANICAL EQUIPMENT SELECTION REPORT")
    title.font = Font(bold=True, size=15, color=_NAVY)
    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 28

    ws.cell(2, 1, f"생성일: {date.today().isoformat()}  |  장비 카테고리: {conditions.get('equipment_category','')}")
    ws.cell(2, 1).font = Font(size=9, italic=True, color="555555")
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 16

    r = 4

    # ── 입력 조건 (가로) ──
    cond_labels = ["카테고리", "설계 압력 (barg)", "설계 온도 (°C)", "요구 유량", "프로젝트 타입"]
    cond_keys   = ["equipment_category", "design_pressure", "design_temperature", "flow_rate_display", "project_type"]
    for i, (lbl, key) in enumerate(zip(cond_labels, cond_keys), 1):
        _hdr(ws, r, i, lbl, span=1)
    r += 1
    for i, key in enumerate(cond_keys, 1):
        val = conditions.get(key, conditions.get(key.replace("_display",""), "—"))
        _cell(ws, r, i, _fmt(val), align="center")
    ws.row_dimensions[r].height = 18
    r += 2

    # ── Top 3 추천 ──
    _hdr(ws, r, 1, "TOP 3 추천 장비", span=9, height=20)
    r += 1

    top3_headers = ["순위", "장비 설명", "프로젝트", "용량", "설계P (barg)", "설계T (°C)", "Vendor", "점수", "리스크"]
    for c, h in enumerate(top3_headers, 1):
        _hdr(ws, r, c, h, fill=_SUB_FILL, font=_SUB_FONT)
    r += 1

    recs = ai_result.get("top_recommendations", [])
    rank_labels = {1: "🥇 1위", 2: "🥈 2위", 3: "🥉 3위"}

    # AI 추천이 있으면 그 순서, 없으면 점수 순으로 Top 3
    top3_rows = []
    if recs:
        for rec in recs[:3]:
            desc = rec.get("equipment_description", "")
            match = comparison_df[
                comparison_df["equipment_description"].fillna("").str.contains(str(desc)[:20], regex=False)
            ]
            row_data = match.iloc[0] if not match.empty else pd.Series()
            top3_rows.append((rec.get("rank", len(top3_rows)+1), rec, row_data))
    else:
        for i, (_, row) in enumerate(comparison_df.head(3).iterrows(), 1):
            top3_rows.append((i, {}, row))

    for rank, rec, row_data in top3_rows:
        risk_row = risk_df[
            risk_df["equipment_description"].fillna("").str.contains(
                str(row_data.get("equipment_description",""))[:20], regex=False
            )
        ] if not risk_df.empty else pd.DataFrame()
        risk_level = risk_row["overall_risk"].iloc[0] if not risk_row.empty else "—"

        cols_data = [
            rank_labels.get(rank, f"#{rank}"),
            _fmt(row_data.get("equipment_description")),
            _fmt(row_data.get("project_name")),
            _fmt(row_data.get("capacity")),
            _fmt(row_data.get("design_pressure_barg")),
            _fmt(row_data.get("design_temperature_degC")),
            _fmt(row_data.get("vendor_name")),
            _fmt(row_data.get("total_score"), 3) if row_data.get("total_score") is not None else "—",
            risk_level,
        ]
        for c, val in enumerate(cols_data, 1):
            f = None
            if c == 8:  # 점수
                try: f = _score_fill(float(val))
                except: pass
            if c == 9:  # 리스크
                f = _RISK_FILL.get(risk_level)
                fnt = _WHITE_FONT if f else None
                _cell(ws, r, c, val, bold=True, align="center", fill=f, font_color="FFFFFF" if f else None)
                ws.cell(r, c).font = Font(bold=True, color="FFFFFF" if f else "000000", size=9)
                continue
            _cell(ws, r, c, val, align="center", fill=f)
        ws.row_dimensions[r].height = 18
        r += 1
    r += 1

    # ── AI 선정 이유 ──
    if recs:
        _hdr(ws, r, 1, "AI 선정 이유", span=9, height=18)
        r += 1
        for rec in recs[:3]:
            rank = rec.get("rank","?")
            reason = rec.get("reason","—")
            concerns = rec.get("concerns","—")
            label = rank_labels.get(rank, f"#{rank}")
            _cell(ws, r, 1, f"{label}  선정 이유", bold=True, align="left")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            reason_cell = _cell(ws, r, 3, reason, wrap=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
            ws.row_dimensions[r].height = 45
            r += 1
            _cell(ws, r, 1, f"{label}  검토 사항", bold=True, align="left")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            _cell(ws, r, 3, concerns, wrap=True)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
            ws.row_dimensions[r].height = 36
            r += 1
        r += 1

    # ── 검토 포인트 ──
    review_pts = ai_result.get("human_review_points", [])
    if review_pts:
        _hdr(ws, r, 1, "반드시 검토할 포인트", span=9, height=18)
        r += 1
        for pt in review_pts:
            c = _cell(ws, r, 1, f"• {pt}", wrap=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
            ws.row_dimensions[r].height = 24
            r += 1
        r += 1

    # ── 대안 제안 ──
    alt = ai_result.get("alternative_suggestions","")
    if alt and alt.strip():
        _hdr(ws, r, 1, "대안 / 설계 변경 제안", span=9, height=18, fill=_SUB_FILL)
        r += 1
        _cell(ws, r, 1, alt, wrap=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        ws.row_dimensions[r].height = 40
        r += 2

    # ── 면책 고지 ──
    disc = ws.cell(r, 1, "⚠ 본 결과는 1차 검토 보조 자료이며, 최종 판단은 담당 엔지니어가 수행해야 합니다.")
    disc.font = Font(italic=True, size=8, color="888888")
    ws.merge_cells(f"A{r}:I{r}")

    _auto_width(ws)
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20


# ─── Sheet 2: Top 10 후보 비교 ───────────────────────────────────────────────

def _sheet_comparison(wb, conditions, comparison_df, risk_df):
    ws = wb.create_sheet("2. Top10 후보 비교")
    if comparison_df.empty:
        ws.cell(1, 1, "데이터 없음")
        return

    ip = conditions.get("design_pressure")
    it = conditions.get("design_temperature")

    headers = [
        "순위", "장비 설명", "프로젝트", "용량",
        "설계P\n(barg)", "P마진\n(%)",
        "설계T\n(°C)", "T마진\n(%)",
        "Vendor", "국가",
        "P-Margin\n점수", "T-Margin\n점수", "용량적합\n점수", "가격\n점수", "납기\n점수", "데이터\n점수",
        "종합점수", "리스크"
    ]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    ws.row_dimensions[1].height = 30

    # 리스크 룩업
    risk_lookup = {}
    if not risk_df.empty:
        for _, row in risk_df.iterrows():
            risk_lookup[str(row.get("equipment_description",""))] = row.get("overall_risk","—")

    score_cols_map = {
        "pressure_margin_score": 11,
        "temperature_margin_score": 12,
        "capacity_fit_score": 13,
        "price_score": 14,
        "lead_time_score": 15,
        "data_completeness_score": 16,
        "total_score": 17,
    }

    for rank_i, (_, row) in enumerate(comparison_df.iterrows(), 1):
        r = rank_i + 1
        dp = row.get("design_pressure_barg")
        dt = row.get("design_temperature_degC")
        pm = _margin_pct(dp, ip)
        tm = _margin_pct(dt, it)
        desc = str(row.get("equipment_description",""))
        risk_level = risk_lookup.get(desc, "—")
        total = row.get("total_score")

        row_vals = [
            rank_i,
            _fmt(row.get("equipment_description")),
            _fmt(row.get("project_name")),
            _fmt(row.get("capacity")),
            _fmt(dp), f"{pm:.0f}%" if pm is not None else "—",
            _fmt(dt), f"{tm:.0f}%" if tm is not None else "—",
            _fmt(row.get("vendor_name")),
            _fmt(row.get("vendor_country")),
        ]

        for c, val in enumerate(row_vals, 1):
            align = "center" if c != 2 else "left"
            _cell(ws, r, c, val, align=align)

        # 점수 컬럼
        for col_key, c in score_cols_map.items():
            val = row.get(col_key)
            f = _score_fill(val) if isinstance(val, float) else None
            _cell(ws, r, c, round(val, 3) if isinstance(val, float) else "—", align="center", fill=f)

        # 리스크
        rf = _RISK_FILL.get(risk_level)
        _cell(ws, r, 18, risk_level, bold=True, align="center", fill=rf)
        if rf:
            ws.cell(r, 18).font = Font(bold=True, color="FFFFFF", size=9)

        ws.row_dimensions[r].height = 16

    _auto_width(ws)
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A2"


# ─── Sheet 3: Top 3 상세 분석 ───────────────────────────────────────────────

def _sheet_top3_detail(wb, conditions, comparison_df, risk_df, ai_result):
    ws = wb.create_sheet("3. Top3 상세 분석")
    if comparison_df.empty:
        ws.cell(1, 1, "데이터 없음")
        return

    top3 = comparison_df.head(3)
    recs = {r.get("equipment_description","")[:30]: r
            for r in ai_result.get("top_recommendations", [])}

    ip = conditions.get("design_pressure")
    it = conditions.get("design_temperature")
    iflow = conditions.get("flow_rate")

    # 컬럼 레이아웃: A=항목명, B=1위, C=2위, D=3위
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B","C","D"]:
        ws.column_dimensions[col_letter].width = 32

    r = 1
    rank_labels = ["🥇 1위", "🥈 2위", "🥉 3위"]

    # 헤더 행
    _hdr(ws, r, 1, "항목")
    for i, lbl in enumerate(rank_labels[:len(top3)], 2):
        _hdr(ws, r, i, lbl)
    ws.row_dimensions[r].height = 22
    r += 1

    def section_hdr(title):
        nonlocal r
        _hdr(ws, r, 1, title, span=4, fill=_SUB_FILL, font=_SUB_FONT, height=18)
        r += 1

    def data_row(label, vals, bold=False, fills=None):
        nonlocal r
        _cell(ws, r, 1, label, bold=bold)
        for i, val in enumerate(vals, 2):
            f = fills[i-2] if fills else None
            _cell(ws, r, i, val, align="center", bold=bold, fill=f)
        ws.row_dimensions[r].height = 16
        r += 1

    rows_list = list(top3.iterrows())

    # ── 기본 정보 ──
    section_hdr("기본 정보")
    data_row("장비 설명", [_fmt(row.get("equipment_description")) for _, row in rows_list])
    data_row("프로젝트명", [_fmt(row.get("project_name")) for _, row in rows_list])
    data_row("프로젝트 타입", [_fmt(row.get("project_type")) for _, row in rows_list])
    data_row("장비 타입", [_fmt(row.get("equipment_type")) for _, row in rows_list])
    data_row("구성", [_fmt(row.get("configuration")) for _, row in rows_list])
    data_row("재질", [_fmt(row.get("material")) for _, row in rows_list])
    data_row("모델 번호", [_fmt(row.get("model_no")) for _, row in rows_list])
    r += 1

    # ── 설계 조건 ──
    section_hdr("설계 조건 vs 입력 조건")
    data_row("용량", [_fmt(row.get("capacity")) for _, row in rows_list], bold=True)
    data_row(f"  입력 유량", [_fmt(iflow) for _ in rows_list])
    data_row("설계 압력 (barg)", [_fmt(row.get("design_pressure_barg")) for _, row in rows_list], bold=True)
    data_row(f"  입력 압력 (barg)", [_fmt(ip) for _ in rows_list])

    # 압력 마진
    p_margins = []
    p_fills = []
    for _, row in rows_list:
        pm = _margin_pct(row.get("design_pressure_barg"), ip)
        p_margins.append(f"{pm:.0f}%" if pm is not None else "—")
        p_fills.append(
            PatternFill("solid", fgColor=_GREEN) if pm is not None and pm >= 20
            else PatternFill("solid", fgColor=_YELLOW) if pm is not None and pm >= 5
            else PatternFill("solid", fgColor=_RED_LIGHT) if pm is not None
            else None
        )
    data_row("  압력 마진 (%)", p_margins, fills=p_fills)

    data_row("설계 온도 (°C)", [_fmt(row.get("design_temperature_degC")) for _, row in rows_list], bold=True)
    data_row(f"  입력 온도 (°C)", [_fmt(it) for _ in rows_list])

    t_margins = []
    t_fills = []
    for _, row in rows_list:
        tm = _margin_pct(row.get("design_temperature_degC"), it)
        t_margins.append(f"{tm:.0f}%" if tm is not None else "—")
        t_fills.append(
            PatternFill("solid", fgColor=_GREEN) if tm is not None and tm >= 20
            else PatternFill("solid", fgColor=_YELLOW) if tm is not None and tm >= 5
            else PatternFill("solid", fgColor=_RED_LIGHT) if tm is not None
            else None
        )
    data_row("  온도 마진 (%)", t_margins, fills=t_fills)
    r += 1

    # ── 치수·중량 ──
    section_hdr("치수 및 중량")
    data_row("L / T-T (mm)", [_fmt(row.get("dim_l_mm")) for _, row in rows_list])
    data_row("W / Dia (mm)", [_fmt(row.get("dim_w_mm")) for _, row in rows_list])
    data_row("H (mm)", [_fmt(row.get("dim_h_mm")) for _, row in rows_list])
    data_row("건조 중량 (kg)", [_fmt(row.get("weight_dry_kg")) for _, row in rows_list])
    data_row("운전 중량 (kg)", [_fmt(row.get("weight_operating_kg")) for _, row in rows_list])
    r += 1

    # ── Vendor 정보 ──
    section_hdr("Vendor 정보")
    data_row("Vendor", [_fmt(row.get("vendor_name")) for _, row in rows_list], bold=True)
    data_row("국가", [_fmt(row.get("vendor_country")) for _, row in rows_list])
    data_row("입찰 가격 (USD)", [_fmt(row.get("price_bidding_usd")) for _, row in rows_list])
    data_row("PO 가격 (USD)", [_fmt(row.get("price_po_usd")) for _, row in rows_list])
    data_row("PO 납기 (주)", [_fmt(row.get("lead_time_po_weeks")) for _, row in rows_list])
    r += 1

    # ── 종합 점수·리스크 ──
    section_hdr("평가 결과")
    score_vals = [_fmt(row.get("total_score"), 3) for _, row in rows_list]
    score_fills = [_score_fill(row.get("total_score")) for _, row in rows_list]
    data_row("종합 점수", score_vals, bold=True, fills=score_fills)

    risk_lookup = {}
    if not risk_df.empty:
        for _, rr in risk_df.iterrows():
            risk_lookup[str(rr.get("equipment_description",""))] = rr.get("overall_risk","—")
    risk_vals = [risk_lookup.get(str(row.get("equipment_description","")), "—") for _, row in rows_list]
    risk_fills = [_RISK_FILL.get(v) for v in risk_vals]
    data_row("종합 리스크", risk_vals, bold=True, fills=risk_fills)
    for col in range(2, 2 + len(top3)):
        cell = ws.cell(r - 1, col)
        if cell.fill and cell.fill.fgColor.rgb != "00000000":
            cell.font = Font(bold=True, color="FFFFFF", size=9)
    r += 1

    # ── AI 분석 ──
    if recs:
        section_hdr("AI 분석")
        for i, (_, row) in enumerate(rows_list):
            desc_key = str(row.get("equipment_description",""))[:30]
            rec = next((v for k, v in recs.items() if k[:20] in desc_key or desc_key[:20] in k), {})
            _cell(ws, r, 1, f"{rank_labels[i]} 선정 이유", bold=True)
            reason_cell = ws.cell(r, 2, rec.get("reason", "AI 추천 데이터 없음"))
            reason_cell.alignment = _WRAP
            reason_cell.border = _BORDER
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            ws.row_dimensions[r].height = 50
            r += 1
            _cell(ws, r, 1, f"{rank_labels[i]} 검토 사항", bold=True)
            concern_cell = ws.cell(r, 2, rec.get("concerns", "—"))
            concern_cell.alignment = _WRAP
            concern_cell.border = _BORDER
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            ws.row_dimensions[r].height = 40
            r += 1


# ─── Sheet 4: 리스크 분석 ────────────────────────────────────────────────────

def _sheet_risk(wb, comparison_df, risk_df):
    ws = wb.create_sheet("4. 리스크 분석")
    if risk_df.empty:
        ws.cell(1, 1, "리스크 데이터 없음")
        return

    r = 1
    # 상단: 장비별 리스크 요약 (가로)
    _hdr(ws, r, 1, "장비별 리스크 요약", span=8, height=20)
    r += 1

    summary_headers = ["장비 설명", "프로젝트", "종합 리스크", "HIGH 건수", "MEDIUM 건수", "LOW 건수"]
    for c, h in enumerate(summary_headers, 1):
        _hdr(ws, r, c, h, fill=_SUB_FILL, font=_SUB_FONT)
    r += 1

    for _, row in risk_df.iterrows():
        level = row.get("overall_risk","—")
        rf = _RISK_FILL.get(level)
        _cell(ws, r, 1, _fmt(row.get("equipment_description")), wrap=True)
        _cell(ws, r, 2, _fmt(row.get("project_name")))
        _cell(ws, r, 3, level, bold=True, align="center", fill=rf)
        if rf:
            ws.cell(r, 3).font = Font(bold=True, color="FFFFFF", size=9)
        _cell(ws, r, 4, row.get("high_count",0), align="center",
              fill=PatternFill("solid", fgColor="FF4444") if row.get("high_count",0) > 0 else None)
        _cell(ws, r, 5, row.get("medium_count",0), align="center",
              fill=PatternFill("solid", fgColor="FFA500") if row.get("medium_count",0) > 0 else None)
        _cell(ws, r, 6, row.get("low_count",0), align="center")
        ws.row_dimensions[r].height = 16
        r += 1
    r += 2

    # 하단: 항목별 상세 (HIGH/MEDIUM만)
    _hdr(ws, r, 1, "HIGH / MEDIUM 리스크 상세", span=8, height=20)
    r += 1
    detail_headers = ["장비 설명", "리스크 항목", "등급", "상세 내용"]
    for c, h in enumerate(detail_headers, 1):
        _hdr(ws, r, c, h, fill=_SUB_FILL, font=_SUB_FONT)
    r += 1

    for _, row in risk_df.iterrows():
        for item in row.get("risk_items",[]):
            if item["level"] not in ("HIGH","MEDIUM"):
                continue
            _cell(ws, r, 1, _fmt(row.get("equipment_description")), wrap=True)
            _cell(ws, r, 2, item["item"])
            rf = _RISK_FILL.get(item["level"])
            _cell(ws, r, 3, item["level"], bold=True, align="center", fill=rf)
            if rf:
                ws.cell(r, 3).font = Font(bold=True, color="FFFFFF", size=9)
            _cell(ws, r, 4, item["detail"], wrap=True)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 28
            r += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["D"].width = 50


# ─── Sheet 5: Vendor 정보 ────────────────────────────────────────────────────

def _sheet_vendor(wb, comparison_df):
    ws = wb.create_sheet("5. Vendor 정보")
    if comparison_df.empty:
        ws.cell(1, 1, "데이터 없음")
        return

    headers = ["순위", "장비 설명", "Vendor", "국가",
               "입찰가 (USD)", "PO가 (USD)", "최종가 (USD)",
               "PO 납기 (주)", "실납기 (주)"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    ws.row_dimensions[1].height = 20

    vendor_cols = ["equipment_description", "vendor_name", "vendor_country",
                   "price_bidding_usd", "price_po_usd", "price_final_usd",
                   "lead_time_po_weeks", "lead_time_actual_weeks"]

    for rank_i, (_, row) in enumerate(comparison_df.iterrows(), 1):
        r = rank_i + 1
        _cell(ws, r, 1, rank_i, align="center")
        for c, col in enumerate(vendor_cols, 2):
            _cell(ws, r, c, _fmt(row.get(col)), align="left" if c == 2 else "center")
        ws.row_dimensions[r].height = 16

    _auto_width(ws)
    ws.column_dimensions["B"].width = 35


# ─── Sheet 6: 전체 후보 목록 ─────────────────────────────────────────────────

_ALL_CANDIDATES_COLS = {
    "equipment_category": "Category",
    "equipment_description": "Description",
    "project_type": "Project Type",
    "client": "Client",
    "built_year": "Built Year",
    "configuration": "Configuration",
    "capacity": "Capacity",
    "model_no": "Model No.",
    "dim_l_mm": "L (mm)",
    "dim_w_mm": "W (mm)",
    "dim_h_mm": "H (mm)",
    "weight_dry_kg": "Dry Wt (kg)",
    "weight_operating_kg": "Op. Wt (kg)",
    "design_pressure_barg": "Design P (barg)",
    "design_temperature_degC": "Design T (°C)",
    "vendor_name": "Vendor",
    "vendor_country": "Country",
}


def _sheet_all_candidates(wb, candidates_df):
    ws = wb.create_sheet("6. 전체 후보 목록")
    if candidates_df.empty:
        ws.cell(1, 1, "데이터 없음")
        return

    cols = [c for c in _ALL_CANDIDATES_COLS if c in candidates_df.columns]
    headers = [_ALL_CANDIDATES_COLS[c] for c in cols]

    for c, h in enumerate(headers, 1):
        _hdr(ws, 1, c, h)
    ws.row_dimensions[1].height = 20

    for r_i, (_, row) in enumerate(candidates_df[cols].iterrows(), 2):
        for c, col in enumerate(cols, 1):
            _cell(ws, r_i, c, _fmt(row.get(col)), align="center" if c > 2 else "left")
        ws.row_dimensions[r_i].height = 15

    _auto_width(ws)
    ws.column_dimensions["B"].width = 35
    ws.freeze_panes = "A2"


# ─── Excel 진입점 ────────────────────────────────────────────────────────────

def _write_excel(path, conditions, candidates_df, comparison_df, risk_df, ai_result):
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_summary(wb, conditions, comparison_df, risk_df, ai_result)
    _sheet_comparison(wb, conditions, comparison_df, risk_df)
    _sheet_top3_detail(wb, conditions, comparison_df, risk_df, ai_result)
    _sheet_risk(wb, comparison_df, risk_df)
    _sheet_vendor(wb, comparison_df)
    _sheet_all_candidates(wb, candidates_df)
    wb.save(path)


# ─── PDF ────────────────────────────────────────────────────────────────────

_PDF_NAVY  = colors.HexColor("#1F3864")
_PDF_BLUE  = colors.HexColor("#2E75B6")
_PDF_GREEN = colors.HexColor("#C6EFCE")
_PDF_RISK  = {"HIGH": colors.HexColor("#FF4444"),
              "MEDIUM": colors.HexColor("#FFA500"),
              "LOW": colors.HexColor("#44BB44")}


def _tbl_style(header_bg=None, row_alt=True):
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), header_bg or _PDF_NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), _KRB),
        ("FONTNAME",   (0, 1), (-1, -1), _KR),
        ("FONTSIZE",   (0, 0), (-1, 0), 8),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
        ("ROWPADDING", (0, 0), (-1, -1), 3),
    ]
    if row_alt:
        style.append(("ROWBACKGROUNDS", (0, 1), (-1, -1),
                       [colors.white, colors.HexColor("#F2F2F2")]))
    return TableStyle(style)


def _write_pdf(path, conditions, comparison_df, risk_df, ai_result):
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    styles = getSampleStyleSheet()
    h1   = ParagraphStyle("h1_kr",   fontSize=14, fontName=_KRB,
                          textColor=_PDF_NAVY, spaceAfter=6)
    h2   = ParagraphStyle("h2_kr",   fontSize=10, fontName=_KRB,
                          textColor=_PDF_BLUE, spaceAfter=4)
    body = ParagraphStyle("body_kr", fontSize=8,  fontName=_KR,
                          leading=13, spaceAfter=3)
    small = ParagraphStyle("small_kr", fontSize=7, fontName=_KR,
                           textColor=colors.grey)
    story = []

    # ── Page 1: 표지 + 입력 조건 ──────────────────────────────────────────
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph("MECHANICAL EQUIPMENT SELECTION REPORT", h1))
    story.append(Paragraph(f"생성일: {date.today().isoformat()}", body))
    story.append(HRFlowable(width="100%", thickness=1, color=_PDF_NAVY))
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("입력 조건", h2))
    cond_data = [["항목", "값"]]
    cond_labels = {
        "equipment_category": "장비 카테고리",
        "design_pressure": "설계 압력 (barg)",
        "design_temperature": "설계 온도 (°C)",
        "flow_rate_display": "요구 유량",
        "project_type": "프로젝트 타입",
    }
    for key, label in cond_labels.items():
        val = conditions.get(key, conditions.get(key.replace("_display",""), "—"))
        cond_data.append([label, _fmt(val)])

    cond_tbl = Table(cond_data, colWidths=[5*cm, 12*cm])
    cond_tbl.setStyle(_tbl_style())
    story.append(cond_tbl)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph(
        "⚠  본 결과는 1차 검토 보조 자료이며, 최종 판단은 담당 엔지니어가 수행해야 합니다.", small
    ))
    story.append(PageBreak())

    # ── Page 2: Top 3 추천 요약 ───────────────────────────────────────────
    story.append(Paragraph("Top 3 추천 장비", h1))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_PDF_BLUE))
    story.append(Spacer(1, 0.3*cm))

    ip = conditions.get("design_pressure")
    it = conditions.get("design_temperature")

    top3 = comparison_df.head(3)
    risk_lookup = {}
    if not risk_df.empty:
        for _, rr in risk_df.iterrows():
            risk_lookup[str(rr.get("equipment_description",""))] = rr.get("overall_risk","—")

    recs = ai_result.get("top_recommendations", [])
    rec_map = {r.get("equipment_description","")[:25]: r for r in recs}

    rank_labels_pdf = ["1위", "2위", "3위"]
    for i, (_, row) in enumerate(top3.iterrows()):
        rank_lbl = rank_labels_pdf[i] if i < 3 else f"#{i+1}"
        desc = str(row.get("equipment_description",""))
        risk_level = risk_lookup.get(desc, "—")
        rec = next((v for k, v in rec_map.items() if k[:15] in desc or desc[:15] in k), {})

        story.append(Paragraph(f"<b>{rank_lbl} — {desc}</b>", h2))

        # 스펙 테이블
        pm = _margin_pct(row.get("design_pressure_barg"), ip)
        tm = _margin_pct(row.get("design_temperature_degC"), it)
        spec_data = [
            ["항목", "장비 스펙", "입력 조건", "마진"],
            ["용량", _fmt(row.get("capacity")), _fmt(conditions.get("flow_rate")), "—"],
            ["설계 압력 (barg)", _fmt(row.get("design_pressure_barg")), _fmt(ip),
             f"{pm:.0f}%" if pm is not None else "—"],
            ["설계 온도 (°C)", _fmt(row.get("design_temperature_degC")), _fmt(it),
             f"{tm:.0f}%" if tm is not None else "—"],
            ["Vendor", _fmt(row.get("vendor_name")), _fmt(row.get("vendor_country")), ""],
            ["종합 점수", _fmt(row.get("total_score"), 3), "", ""],
            ["리스크 등급", risk_level, "", ""],
        ]
        spec_tbl = Table(spec_data, colWidths=[4*cm, 5*cm, 4*cm, 4*cm])
        style = _tbl_style()
        # 리스크 셀 색상
        risk_color = _PDF_RISK.get(risk_level)
        if risk_color:
            style.add("BACKGROUND", (1, 6), (1, 6), risk_color)
            style.add("TEXTCOLOR",  (1, 6), (1, 6), colors.white)
        spec_tbl.setStyle(style)
        story.append(spec_tbl)

        # AI 이유
        if rec.get("reason"):
            story.append(Spacer(1, 0.2*cm))
            story.append(Paragraph(f"<b>선정 이유:</b> {rec.get('reason','')}", body))
        if rec.get("concerns"):
            story.append(Paragraph(f"<b>검토 사항:</b> {rec.get('concerns','')}", body))

        story.append(Spacer(1, 0.4*cm))

    story.append(PageBreak())

    # ── Page 3: 리스크 요약 + 검토 포인트 ────────────────────────────────
    story.append(Paragraph("리스크 분석 요약", h1))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_PDF_BLUE))
    story.append(Spacer(1, 0.3*cm))

    if not risk_df.empty:
        risk_data = [["장비 설명", "종합 리스크", "HIGH", "MED", "LOW"]]
        for _, row in risk_df.iterrows():
            risk_data.append([
                str(row.get("equipment_description",""))[:45],
                row.get("overall_risk","—"),
                str(row.get("high_count",0)),
                str(row.get("medium_count",0)),
                str(row.get("low_count",0)),
            ])
        risk_tbl = Table(risk_data, colWidths=[9*cm, 3*cm, 1.5*cm, 1.5*cm, 1.5*cm])
        rs = _tbl_style()
        for ri, rrow in enumerate(risk_df.itertuples(), 1):
            rc = _PDF_RISK.get(rrow.overall_risk)
            if rc:
                rs.add("BACKGROUND", (1, ri), (1, ri), rc)
                rs.add("TEXTCOLOR",  (1, ri), (1, ri), colors.white)
        risk_tbl.setStyle(rs)
        story.append(risk_tbl)
        story.append(Spacer(1, 0.5*cm))

    review_pts = ai_result.get("human_review_points", [])
    if review_pts:
        story.append(Paragraph("반드시 검토할 포인트", h2))
        for pt in review_pts:
            story.append(Paragraph(f"• {pt}", body))
        story.append(Spacer(1, 0.4*cm))

    alt = ai_result.get("alternative_suggestions","")
    if alt and alt.strip():
        story.append(Paragraph("대안 / 설계 변경 제안", h2))
        story.append(Paragraph(alt, body))
        story.append(Spacer(1, 0.4*cm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Paragraph(
        "⚠  본 결과는 1차 검토 보조 자료이며, 최종 판단은 담당 엔지니어가 수행해야 합니다. "
        "AI 추천 및 점수 산정 결과는 과거 DB 기준이며, 현행 프로젝트 요구사항·규격·최신 Vendor 사양과 "
        "반드시 대조 확인하십시오.", small
    ))

    doc.build(story)
